from docx import Document
import openpyxl
import re
from openpyxl.styles import Alignment

# Открываем документ DOCX
doc = Document("your_document.docx")

# Создаем новый файл Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки для Excel
column_headers = {
    'A1': 'Пациент',
    'B1': 'Год рождения',
    'C1': 'Дата исследования',
    'D1': 'Правая почка\n(размеры и толщина паренхимы)',
    'E1': 'Левая почка\n(размеры и толщина паренхимы)',
    'F1': 'Правая почка\n(чашечно-лоханочная система)',
    'G1': 'Левая почка\n(чашечно-лоханочная система)',
    'H1': 'Правая лоханка\n(размер)',
    'I1': 'Левая лоханка\n(размер)',
    'J1': 'Правая почка\n(конкремент размер)',
    'K1': 'Левая почка\n(конкремент размер)',
    'L1': 'Правая почка\n(паренхима толщина)',
    'M1': 'Левая почка\n(паренхима толщина)',
    'N1': 'Правая почка\n(высота)',
    'O1': 'Правая почка\n(ширина)',
    'P1': 'Левая почка\n(высота)',
    'Q1': 'Левая почка\n(ширина)',
    'R1': 'Предстательная железа\n(ширина)',
    'S1': 'Предстательная железа\n(высота)',
    'T1': 'Предстательная железа\n(передне-задний размер)',
    'U1': 'объем железы',
}

#Добавим заголовки столбцов на лист и установите перенос текста
for cell, header_text in column_headers.items():
    worksheet[cell] = header_text
    worksheet.column_dimensions[cell[:-1]].width = 20  # Отрегулируем ширину столбца по мере необходимости
    worksheet[cell].alignment = Alignment(wrapText=True)

# Устанавливаем ширину столбцов
column_widths = {
    'A': 40,
    'B': 10,
    'C': 15,
    'D': 30,
    'E': 30,
    'F': 30,
    'G': 30,
    'H': 20,
    'I': 20,
    'J': 20,
    'K': 20,
    'L': 20,
    'M': 20,
    'N': 20,
    'O': 20,
    'P': 20,
    'Q': 20,
    'R': 20,
    'S': 20,
    'T': 20,
    'U': 20,
}

#Устанавливаем заголовки и ширину столбцов
for cell, header in column_headers.items():
    worksheet[cell] = header

for column, width in column_widths.items():
    worksheet.column_dimensions[column].width = width

worksheet.row_dimensions[1].height = 30

# Индекс строки в Excel
row_index = 2

# Переменные для хранения информации о пациенте
patient_name = None
patient_birth_year = None
examination_date = None

# Регулярное выражение для удаления ненужных текстов
pattern = r'(не увеличена, расположена в типичном месте|обычной акустической плотности, кортико-медуллярная дифференциация сохранена)'


# Проверка наличия слова в тексте
def check_word_in_text(word, text):
    return True if text.find(word) != -1 else False


for paragraph in doc.paragraphs:
    text = paragraph.text
    if "Пациент: " in text:
        patient_name = text.replace("Пациент: ", "").strip()
    # if "Год рождения:" in text:
    #     clean_text = re.sub(r'\D', '', text)  # Извлекаем только цифры из строки
    #     try:
    #         patient_birth_year = int(clean_text)
    #     except ValueError:
    #         print(f"Ошибка: невозможно преобразовать '{clean_text}' к целому числу года рождения.")
    elif "Год рождения:" in text:
        patient_birth_year = text.replace("Год рождения:", "").strip()
        try:
            patient_birth_year = int(patient_birth_year)
        except ValueError:
            print(f"Ошибка: невозможно преобразовать '{patient_birth_year}' к целому числу года рождения.")
    elif "Дата исследования:" in text:
        examination_date = text.replace("Дата исследования:", "").strip()

    if "Правая почка" in text and "Левая почка" in text:
        right_kidney = text.split("Правая почка")[1].split("Левая почка")[0].strip()
        left_kidney = text.split("Левая почка")[1].strip()

        # Извлекаем размеры и толщину паренхимы для правой почки
        right_kidney_data = re.sub(pattern, '', right_kidney).strip()
        right_kidney_size = re.search(r'(\d+х\d+)', right_kidney_data).group(1)

        # Извлекаем размеры и толщину паренхимы для левой почки
        left_kidney_data = re.sub(pattern, '', left_kidney).strip()
        left_kidney_size = re.search(r'(\d+х\d+)', left_kidney_data).group(1)

        right_kidney_sizes_match = re.search(r'(\d+)х(\d+)', right_kidney_data)
        if right_kidney_sizes_match:
            right_kidney_height = int(right_kidney_sizes_match.group(1))
            right_kidney_width = int(right_kidney_sizes_match.group(2))

        left_kidney_sizes_match = re.search(r'(\d+)х(\d+)', left_kidney_data)
        if left_kidney_sizes_match:
            left_kidney_height = int(left_kidney_sizes_match.group(1))
            left_kidney_width = int(left_kidney_sizes_match.group(2))
        else:
            left_kidney_height = None
            left_kidney_width = None

        # Извлекаем информацию о чашечно-лоханочной системе
        right_kidney_system = None
        left_kidney_system = None

        if "чашечно-лоханочная система не расширена" not in right_kidney:
            right_kidney_system = "расширена"
        else:
            right_kidney_system = "не расширена"

        if "чашечно-лоханочная система не расширена" not in left_kidney:
            left_kidney_system = "расширена"
        else:
            left_kidney_system = "не расширена"
        worksheet.cell(row=row_index, column=6, value=right_kidney_system)
        worksheet.cell(row=row_index, column=7, value=left_kidney_system)

        right_kidney_lohanka = None
        left_kidney_lohanka = None

        if right_kidney_system == "расширена":
            right_kidney_lohanka_match = re.search(r'лоханка (\d+) мм', right_kidney)
            if right_kidney_lohanka_match:
                right_kidney_lohanka = int(right_kidney_lohanka_match.group(1))

        if left_kidney_system == "расширена":
            left_kidney_lohanka_match = re.search(r'лоханка (\d+) мм', left_kidney)
            if left_kidney_lohanka_match:
                left_kidney_lohanka = int(left_kidney_lohanka_match.group(1))

        # Извлекаем информацию о конкрементах для правой почки
        right_kidney_concrement = None
        right_kidney_concrement_match = re.search(r'конкремент размером (\d+) мм', right_kidney)
        if right_kidney_concrement_match:
            right_kidney_concrement = int(right_kidney_concrement_match.group(1))

        # Извлекаем информацию о конкрементах для левой почки
        left_kidney_concrement = None
        left_kidney_concrement_match = re.search(r'конкремент размером (\d+) мм', left_kidney)
        if left_kidney_concrement_match:
            left_kidney_concrement = int(left_kidney_concrement_match.group(1))

        # Извлекаем информацию о паренхиме почек
        right_kidney_parenchyma = None
        left_kidney_parenchyma = None

        right_kidney_parenchyma_match = re.search(r'паренхима толщиной (\d+) мм', right_kidney)
        if right_kidney_parenchyma_match:
            right_kidney_parenchyma = int(right_kidney_parenchyma_match.group(1))

            left_kidney_parenchyma_match = re.search(r'паренхима толщиной (\d+) мм', left_kidney)
            if left_kidney_parenchyma_match:
                left_kidney_parenchyma = int(left_kidney_parenchyma_match.group(1))
        else:
            pass

        prostate_width = None
        prostate_height = None
        prostate_depth = None

        # Извлекаем данные о предстательной железе
        if "Предстательная железа" in text:
            prostate_data = re.search(r'Предстательная железа\s+(\d+)х(\d+)х(\d+) мм', text)
            if prostate_data:
                prostate_width = int(prostate_data.group(1))
                prostate_height = int(prostate_data.group(2))
                prostate_depth = int(prostate_data.group(3))

            worksheet.cell(row=row_index, column=18, value=prostate_width)
            worksheet.cell(row=row_index, column=19, value=prostate_height)
            worksheet.cell(row=row_index, column=20, value=prostate_depth)
        else:
            # Пропустить добавление данных о предстательной железе, оставив ячейки пустыми
            worksheet.cell(row=row_index, column=18, value=None)
            worksheet.cell(row=row_index, column=19, value=None)
            worksheet.cell(row=row_index, column=20, value=None)

        # Извлекаем данные о предстательной железе
        prostate_volume = None
        prostate_match = re.search(r'объем железы (\d+) мл', text)
        if prostate_match:
            prostate_volume = int(prostate_match.group(1))
            worksheet.cell(row=row_index, column=21, value=prostate_volume)
        else:
            worksheet.cell(row=row_index, column=21, value=None)

            # Добавляем данные в Excel
            worksheet.cell(row=row_index, column=1, value=patient_name)
            worksheet.cell(row=row_index, column=2, value=patient_birth_year)
            worksheet.cell(row=row_index, column=3, value=examination_date)
            worksheet.cell(row=row_index, column=4, value=right_kidney_size)
            worksheet.cell(row=row_index, column=5, value=left_kidney_size)

            worksheet.cell(row=row_index, column=8, value=right_kidney_lohanka)
            worksheet.cell(row=row_index, column=9, value=left_kidney_lohanka)
            worksheet.cell(row=row_index, column=10, value=right_kidney_concrement)
            worksheet.cell(row=row_index, column=11, value=left_kidney_concrement)
            worksheet.cell(row=row_index, column=12, value=right_kidney_parenchyma)
            worksheet.cell(row=row_index, column=13, value=left_kidney_parenchyma)

            worksheet.cell(row=row_index, column=14, value=right_kidney_height)
            worksheet.cell(row=row_index, column=15, value=right_kidney_width)
            worksheet.cell(row=row_index, column=16, value=left_kidney_height)
            worksheet.cell(row=row_index, column=17, value=left_kidney_width)

            row_index += 1

    # Сохраняем результаты в файл Excel
workbook.save("result.xlsx")
print("Результат сохранен в файл 'result.xlsx'")
