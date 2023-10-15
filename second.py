from docx import Document
import openpyxl
import re
from openpyxl.styles import Alignment, Font

# Открываем документ DOCX
doc = Document("your_document.docx")

# Создаем новый файл Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Индекс строки в Excel
row_index = 2

# Устанавливаем значения по умолчанию
patient_name = ""
patient_birth_year = 0
examination_date = ""
right_kidney = ""
left_kidney = ""
right_kidney_size = None  # Установка значения по умолчанию
right_kidney_lohanka = 0
left_kidney_lohanka = 0
right_kidney_system = ""
left_kidney_system = ""
word_found = False


# Заголовки для Excel
column_headers = {
    'A1': 'Пациент',
    'B1': 'Год рождения',
    'C1': 'Дата исследования',
    'D1': 'Правая почка\n(размеры и толщина паренхимы)',
    'E1': 'Левая почка\n(размеры и толщина паренхимы)',
    'F1': 'Правая лоханка',
    'G1': 'Левая лоханка',
    'H1': 'Правая лоханка\n(размер)',
    'I1': 'Левая лоханка\n(размер)',
    'J1': 'Правая почка\n(конкремент размер)',
    'K1': 'Левая почка\n(конкремент размер)',
    'L1': 'Правая почка\n(паренхима толщина)',
    'M1': 'Левая почка\n(паренхима толщина)',
    'N1': 'Правая почка\n(высота)',
    'O1': 'Правая почка\n(ширина)',
    'P1': 'Левая почка\n(высота)',
    'Q1': 'Левая почка\n(ширина)',
    'R1': 'Предстательная железа\n(ширина)',
    'S1': 'Предстательная железа\n(высота)',
    'T1': 'Предстательная железа\n(передне-задний размер)',
    'U1': 'объем железы',
    'V1': 'киста правого эякуляторного \nпротока размером',
    'W1': 'киста левого эякуляторного \nпротока размером',
    'X1': 'Яичко правое ',
    'Y1': 'Правая\n головка придатка',
    'Z1': 'Правая\n придатка киста',
    'AA1': 'Яичко левое',
    'AB1': 'Левая\n головка придатка',
    'AC1': 'Левая\n придатка киста',
    'AD1': 'Вены гроздевидного\n сплетения',
    'AE1': 'пробы Вальсальвы ',
    'AF1': 'Эхопризнаки',
}

#Добавим заголовки столбцов на лист и установим перенос текста
for cell, header_text in column_headers.items():
    worksheet[cell] = header_text
    worksheet.column_dimensions[cell[:-1]].width = 20  # Отрегулируем ширину столбца по мере необходимости
    worksheet[cell].alignment = Alignment(wrapText=True)
    worksheet[cell].alignment = Alignment(wrap_text=True, horizontal='center')
    worksheet[cell].font = Font(bold=True)

column_widths = {
    'A': 30,
    'B': 10,
    'C': 15,
    'D': 30,
    'E': 30,
    'F': 30,
    'G': 30,
    'H': 15,
    'I': 15,
    'J': 20,
    'K': 20,
    'L': 20,
    'M': 20,
    'N': 20,
    'O': 20,
    'P': 20,
    'Q': 20,
    'R': 20,
    'S': 20,
    'T': 20,
    'U': 20,
    'V': 20,
    'W': 20,
    'X': 20,
    'Y': 22,
    'Z': 15,
    'AA': 20,
    'AB': 22,
    'AC': 15,
    'AD': 15,
    'AE': 20,
    'AF': 200,
}

#Устанавливаем заголовки и ширину столбцов
for cell, header in column_headers.items():
    worksheet[cell] = header

for column, width in column_widths.items():
    worksheet.column_dimensions[column].width = width

worksheet.row_dimensions[1].height = 50

# Регулярное выражение для удаления ненужных текстов
pattern = r'(не увеличена, расположена в типичном месте|обычной акустической плотности, кортико-медуллярная дифференциация сохранена Эхопризнаки)'

for paragraph in doc.paragraphs:
    text = paragraph.text
    if "Пациент: " in text:
        patient_name = text.replace("Пациент: ", "").strip()
    # if "Эхопризнаки " in text:
    #     eho_data = text.replace("Эхопризнаки: ", "").strip()
    #     print(eho_data)
    if "Год рождения:" in text:
        clean_text = text[-4:]  # Извлекаем последние четыре символа, предполагая, что это год рождения
        try:
            patient_birth_year = int(clean_text)
        except ValueError:
            print(f"Ошибка: невозможно преобразовать '{clean_text}' к целому числу года рождения.")
            patient_birth_year = 0  # Или другое значение по умолчанию
    elif "Дата исследования:" in text:
        examination_date = text.replace("Дата исследования:", "").strip()

        # Добавляем данные в Excel паспортные части
        worksheet.cell(row=row_index, column=1, value=patient_name)
        worksheet.cell(row=row_index, column=2, value=patient_birth_year)
        worksheet.cell(row=row_index, column=3, value=examination_date)
        row_index += 1

    elif "Правая почка" in text and "Левая почка" in text:
        right_kidney = text.split("Правая почка")[1].split("Левая почка")[0].strip()
        left_kidney = text.split("Левая почка")[1].strip()

        # Извлекаем размеры и толщину паренхимы для правой почки
        right_kidney_data = re.sub(pattern, '', right_kidney).strip()
        if re.search(r'(\d+х\d+)', right_kidney_data):
            right_kidney_size = re.search(r'(\d+х\d+)', right_kidney_data).group(1)

        # Извлекаем размеры и толщину паренхимы для левой почки
        left_kidney_data = re.sub(pattern, '', left_kidney).strip()
        if re.search(r'(\d+х\d+)', left_kidney_data):
            left_kidney_size = re.search(r'(\d+х\d+)', left_kidney_data).group(1)
        else:
            left_kidney_size = None

        right_kidney_sizes_match = re.search(r'(\d+)х(\d+)', right_kidney_data)
        if right_kidney_sizes_match:
            right_kidney_height = int(right_kidney_sizes_match.group(1))
            right_kidney_width = int(right_kidney_sizes_match.group(2))

        left_kidney_sizes_match = re.search(r'(\d+)х(\d+)', left_kidney_data)
        if left_kidney_sizes_match:
            left_kidney_height = int(left_kidney_sizes_match.group(1))
            left_kidney_width = int(left_kidney_sizes_match.group(2))
        else:
            left_kidney_height = None
            left_kidney_width = None

        # Добавляем данные в Excel размеры и толщину паренхимы почки
        worksheet.cell(row=row_index, column=4, value=right_kidney_size if "Правая почка" in text else None)
        worksheet.cell(row=row_index, column=5, value=left_kidney_size if "Левая почка" in text else None)


        # Извлекаем информацию о чашечно-лоханочной системе
        right_kidney_system = None
        left_kidney_system = None

        if "чашечно-лоханочная система не расширена" not in right_kidney:
            right_kidney_system = "расширена"
        else:
            right_kidney_system = "не расширена"

        if "чашечно-лоханочная система не расширена" not in left_kidney:
            left_kidney_system = "расширена"
        else:
            left_kidney_system = "не расширена"
        worksheet.cell(row=row_index, column=6, value=right_kidney_system)
        worksheet.cell(row=row_index, column=7, value=left_kidney_system)

        right_kidney_lohanka = None
        left_kidney_lohanka = None

        if right_kidney_system == "расширена":
            right_kidney_lohanka_match = re.search(r'лоханка (\d+) мм', right_kidney)
            if right_kidney_lohanka_match:
                right_kidney_lohanka = int(right_kidney_lohanka_match.group(1))

        if left_kidney_system == "расширена":
            left_kidney_lohanka_match = re.search(r'лоханка (\d+) мм', left_kidney)
            if left_kidney_lohanka_match:
                left_kidney_lohanka = int(left_kidney_lohanka_match.group(1))

        # Извлекаем информацию о конкрементах для правой почки
        right_kidney_concrement = None
        right_kidney_concrement_match = re.search(r'конкремент размером (\d+) мм', right_kidney)
        if right_kidney_concrement_match:
            right_kidney_concrement = int(right_kidney_concrement_match.group(1))

        # Извлекаем информацию о конкрементах для левой почки
        left_kidney_concrement = None
        left_kidney_concrement_match = re.search(r'конкремент размером (\d+) мм', left_kidney)
        if left_kidney_concrement_match:
            left_kidney_concrement = int(left_kidney_concrement_match.group(1))

        # Извлекаем информацию о паренхиме почек
        right_kidney_parenchyma = None
        left_kidney_parenchyma = None

        right_kidney_parenchyma_match = re.search(r'паренхима толщиной (\d+) мм', right_kidney)
        if right_kidney_parenchyma_match:
            right_kidney_parenchyma = int(right_kidney_parenchyma_match.group(1))

            left_kidney_parenchyma_match = re.search(r'паренхима толщиной (\d+) мм', left_kidney)
            if left_kidney_parenchyma_match:
                left_kidney_parenchyma = int(left_kidney_parenchyma_match.group(1))
        else:
            pass

        prostate_width = None
        prostate_height = None
        prostate_depth = None

        # Извлекаем данные о предстательной железе
        if "Предстательная железа" in text:
            prostate_data = re.search(r'Предстательная железа\s+(\d+)х(\d+)х(\d+) мм', text)
            if prostate_data:
                prostate_width = int(prostate_data.group(1))
                prostate_height = int(prostate_data.group(2))
                prostate_depth = int(prostate_data.group(3))

            worksheet.cell(row=row_index, column=18, value=prostate_width)
            worksheet.cell(row=row_index, column=19, value=prostate_height)
            worksheet.cell(row=row_index, column=20, value=prostate_depth)
        else:
            # Пропустить добавление данных о предстательной железе, оставив ячейки пустыми
            worksheet.cell(row=row_index, column=18, value=None)
            worksheet.cell(row=row_index, column=19, value=None)
            worksheet.cell(row=row_index, column=20, value=None)

        # Извлекаем данные о предстательной железе
        prostate_volume = None
        prostate_match = re.search(r'объем железы (\d+) мл', text)
        if prostate_match:
            prostate_volume = int(prostate_match.group(1))
            worksheet.cell(row=row_index, column=21, value=prostate_volume)
        else:
            worksheet.cell(row=row_index, column=21, value=None)

        worksheet.cell(row=row_index, column=8, value=right_kidney_lohanka)
        worksheet.cell(row=row_index, column=9, value=left_kidney_lohanka)
        worksheet.cell(row=row_index, column=10, value=right_kidney_concrement)
        worksheet.cell(row=row_index, column=11, value=left_kidney_concrement)
        worksheet.cell(row=row_index, column=12, value=right_kidney_parenchyma)
        worksheet.cell(row=row_index, column=13, value=left_kidney_parenchyma)
        worksheet.cell(row=row_index, column=14, value=right_kidney_height)
        worksheet.cell(row=row_index, column=15, value=right_kidney_width)
        worksheet.cell(row=row_index, column=16, value=left_kidney_height)
        worksheet.cell(row=row_index, column=17, value=left_kidney_width)

        # Извлекаем информацию о кистах эякуляторного протока
        right_cyst_size = None
        left_cyst_size = None

        cyst_size_match = re.search(r'киста\s+эякуляторного\s+протока\s+размером\s+(\d+)\s+мм', text)
        if cyst_size_match:
            cyst_size = int(cyst_size_match.group(1))
            if "правого" in text:
                right_cyst_size = cyst_size
                print(f"Размер правого кисты эякуляторного протока: {right_cyst_size} мм")
                worksheet.cell(row=row_index, column=22,
                               value=right_cyst_size)  # Добавление значения размера правой кисты эякуляторного протока
            elif "левого" in text:
                left_cyst_size = cyst_size
                print(f"Размер левого кисты эякуляторного протока: {left_cyst_size} мм")
                worksheet.cell(row=row_index, column=23,
                               value=left_cyst_size)  # Добавление значения размера левой кисты эякуляторного протока

        # Инициализация переменных для данных о правом яичке
        right_testicle_size = None
        right_appendix_head_size = None
        right_appendix_cyst_size = None

        # Инициализация переменных для данных о левом яичке
        left_testicle_size = None
        left_appendix_head_size = None
        left_appendix_cyst_size = None

        # Инициализация переменных для данных о венах гроздевидного сплетения семенного канатика
        semen_cord_vein_size_left = None

        # Инициализация переменной для данных о пробе Вальсальвы
        valsalva_test = None

        # Регулярное выражение для извлечения значений размеров
        size_pattern = r'(\d+(\.\d+)?)х(\d+(\.\d+)?) мм'

        # Извлечение данных о правом яичке
        right_testicle_match = re.search(r'Яичко правое\s+' + size_pattern, text)
        if right_testicle_match:
            right_testicle_size = right_testicle_match.group()

        right_appendix_head_match = re.search(r'головка придатка\s+' + size_pattern, text)
        if right_appendix_head_match:
            right_appendix_head_size = right_appendix_head_match.group()

        right_appendix_cyst_match = re.search(r'головки придатка яичка лоцируется киста\s+размером\s+(\d+(\.\d+)?) мм',
                                              text)
        if right_appendix_cyst_match:
            right_appendix_cyst_size = float(right_appendix_cyst_match.group(1))

        # Извлечение данных о левом яичке
        left_testicle_match = re.search(r'Яичко левое\s+' + size_pattern, text)
        if left_testicle_match:
            left_testicle_size = left_testicle_match.group()

        left_appendix_head_match = re.search(r'головка придатка\s+' + size_pattern, text)
        if left_appendix_head_match:
            left_appendix_head_size = left_appendix_head_match.group()

        left_appendix_cyst_match = re.search(r'головки придатка яичка лоцируется киста\s+размером\s+(\d+(\.\d+)?) мм',
                                             text)
        if left_appendix_cyst_match:
            left_appendix_cyst_size = left_appendix_cyst_match.group(1)

        # Извлечение данных о венах гроздевидного сплетения семенного канатика
        semen_cord_vein_match_left = re.search(
            r'Вены гроздевидного сплетения семенного канатика слева расширены до\s+(\d+(\.\d+)?) мм', text)
        if semen_cord_vein_match_left:
            semen_cord_vein_size_left = float(semen_cord_vein_match_left.group(1))

        # Извлечение данных о пробе Вальсальвы
        if 'проба Вальсальвы положительная' in text:
            valsalva_test = 'положительная'

        # Добавление извлеченных данных в соответствующие ячейки Excel
        worksheet.cell(row=row_index, column=24, value=right_testicle_size)
        worksheet.cell(row=row_index, column=25, value=right_appendix_head_size)
        worksheet.cell(row=row_index, column=26, value=right_appendix_cyst_size)
        worksheet.cell(row=row_index, column=27, value=left_testicle_size)
        worksheet.cell(row=row_index, column=28, value=left_appendix_head_size)
        worksheet.cell(row=row_index, column=29, value=left_appendix_cyst_size)
        worksheet.cell(row=row_index, column=30, value=semen_cord_vein_size_left)
        worksheet.cell(row=row_index, column=31, value=valsalva_test)

    if "Эхопризнаки " in text:
        eho_data = text.replace("Эхопризнаки: ", "").strip()
        print(eho_data)
        worksheet.cell(row=row_index, column=32, value=eho_data)


    # eho_text = None
    #
    # eho_text_match = re.search(r'Эхопризнаки (.*)', text)
    # if eho_text_match:
    #     eho_text = eho_text_match.group(1)
    #
    # if eho_text:
    #     print(eho_text)
    # else:
    #     print("Подстрока 'Эхопризнаки ' не найдена")
    #     worksheet.cell(row=row_index, column=32, value=eho_text_match)


# Сохраняем файл Excel
workbook.save("output.xlsx")


